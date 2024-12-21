import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
import csv

# 确保文件路径正确
file_path = os.path.join(os.getcwd(), 'xlsx/2024.8.xlsx')

# 加载 Excel 文件
workbook = load_workbook(file_path)

# 定义黄色、蓝色和橙色填充的 RGB 颜色值
yellow_rgb = 'FFF5C400'  # ARGB 格式
blue_rgb = 'FF8CDDFA'  # ARGB 格式
orange_rgb = 'FFF88825'  # ARGB 格式

# 遍历所有工作表
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    print(f"正在处理工作表: {sheet_name}")

    # 取消第一列的单元格合并，并用原内容填充 I 列的每个单元格
    for merged_cell in list(sheet.merged_cells.ranges):
        if merged_cell.min_col == 1:
            top_left_cell_value = sheet.cell(merged_cell.min_row, merged_cell.min_col).value
            sheet.unmerge_cells(str(merged_cell))
            for row in range(merged_cell.min_row, merged_cell.max_row + 1):
                sheet.cell(row, 9).value = top_left_cell_value  # I 列对应的列号是 9

    # 收集需要删除的行号
    rows_to_delete = set()

    # 遍历所有单元格
    for row in sheet.iter_rows():
        for cell in row:
            fill = cell.fill
            if fill and fill.fgColor and (fill.fgColor.rgb == yellow_rgb or fill.fgColor.rgb == blue_rgb or fill.fgColor.rgb == orange_rgb):
                rows_to_delete.add(cell.row)

    # 删除收集到的行的 C-H 列
    for row in rows_to_delete:
        for col in range(3, 9):  # C-H 列对应的列号是 3-8
            sheet.cell(row=row, column=col).value = None

    # 删除标题行
    sheet.delete_rows(1, 1)

    # 在 I 列第一行添加“日期”
    sheet.insert_rows(1)
    sheet.cell(1, 9).value = "日期"

    # 删除 A 和 B 列
    sheet.delete_cols(1, 2)

    # 删除 F 列
    sheet.delete_cols(6, 1)

    # 删除符合 RGB 颜色的行
    for row in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row, 1)

    # 取消所有单元格的边框线
    no_border = Border(left=Side(border_style=None),
                       right=Side(border_style=None),
                       top=Side(border_style=None),
                       bottom=Side(border_style=None))

    # 设置所有单元格的字体为宋体
    font = Font(name='宋体')

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = no_border
            cell.font = font

    # 删除整个文档的最后一行
    sheet.delete_rows(sheet.max_row, 1)

    # 删除第二行
    sheet.delete_rows(2, 1)

# 确保目标目录存在
output_dir = 'xlsx/converted'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# 保存修改后的文件到指定路径
output_path = os.path.join(output_dir, 'converted.xlsx')
workbook.save(output_path)

# 加载修改后的 Excel 文件
workbook = load_workbook(output_path)

# 确保 CSV 输出目录存在
csv_output_dir = os.path.join(output_dir, 'csv')
if not os.path.exists(csv_output_dir):
    os.makedirs(csv_output_dir)

# 遍历每个工作表并转换为 CSV
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    csv_file_path = os.path.join(csv_output_dir, f"{sheet_name}.csv")
    with open(csv_file_path, mode='w', newline='', encoding='utf-8-sig') as csv_file:
        csv_writer = csv.writer(csv_file)
        for row in sheet.iter_rows(values_only=True):
            csv_writer.writerow(row)

print("所有工作表已成功转换为 CSV 文件。")